from fastapi import FastAPI, HTTPException, Query, Body, Path
from pydantic import BaseModel
from typing import Dict, List
import pandas as pd
import openpyxl
from fastapi.middleware.cors import CORSMiddleware
from models import (
    ReceptionRameModel,
    EnvoiRameModel,
    ReceptionPrestataireRLAModel,
    EnvoiPrestataireRLAModel,
    ReparationModuleModel,
    EssaiModel
)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Ou spécifiez une liste des domaines autorisés
    allow_credentials=True,
    allow_methods=["*"],  # Permet tous les types de requêtes HTTP (GET, POST, etc.)
    allow_headers=["*"],  # Permet tous les en-têtes (par exemple Authorization)
)

# Chemin du fichier Excel sur votre PC
EXCEL_FILE_PATH = '/home/youssef/Téléchargements/Gestion.xlsm'  # Remplacez par le chemin réel

# Fonction pour lire les données de la feuille de calcul spécifiée depuis le fichier Excel
def read_excel_data(sheet_name: str):
    try:
        # Charger le fichier Excel et lire la feuille de calcul spécifiée
        df = pd.read_excel(EXCEL_FILE_PATH, sheet_name=sheet_name, engine='openpyxl')
        
        # Remplacer les valeurs NaN par une chaîne vide
        df = df.fillna("")
        
        return df
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Fichier Excel introuvable")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Feuille de calcul non trouvée ou erreur de valeur: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la lecture du fichier Excel: {str(e)}")

# Fonction pour écrire les données modifiées dans le fichier Excel
def write_excel_data(sheet_name: str, df: pd.DataFrame):
    try:
        # Charger le classeur Excel existant
        with pd.ExcelWriter(EXCEL_FILE_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Supprimer l'ancienne feuille si elle existe
            book = writer.book
            if sheet_name in book.sheetnames:
                std = book[sheet_name]
                book.remove(std)
            
            # Écrire les nouvelles données
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Fichier Excel introuvable")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'écriture dans le fichier Excel: {str(e)}")

# Endpoint pour récupérer les colonnes d'une feuille Excel
@app.get("/api/v1/excel_columns")
def get_excel_columns(sheet_name: str = Query(..., description="Nom de la feuille de calcul à lire")):
    try:
        # Lire la première ligne pour obtenir les colonnes
        df = pd.read_excel(EXCEL_FILE_PATH, sheet_name=sheet_name, engine='openpyxl', nrows=0)
        return {"columns": list(df.columns)}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Fichier Excel introuvable")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Feuille de calcul non trouvée ou erreur de valeur: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la lecture du fichier Excel: {str(e)}")

# Endpoint pour récupérer les données d'une feuille Excel
@app.get("/api/v1/db_data")
def get_db_data(sheet_name: str = Query(..., description="Nom de la feuille de calcul à lire")):
    df = read_excel_data(sheet_name)

    if df.empty:
        raise HTTPException(status_code=404, detail="Aucune donnée trouvée dans la feuille de calcul spécifiée")

    # Ajouter l'index comme colonne "Index"
    df.reset_index(inplace=True)

    # Convertir le DataFrame en une liste de dictionnaires
    data = df.to_dict(orient='records')

    return {"db_data": data}

# Endpoint pour supprimer une ligne
@app.delete("/api/v1/db_data/{index}")
def delete_row(sheet_name: str = Query(..., description="Nom de la feuille de calcul"), index: int = Path(..., description="Index de la ligne à supprimer")):
    df = read_excel_data(sheet_name)
    
    # Vérifier que l'index est valide
    if index < 0 or index >= len(df):
        raise HTTPException(status_code=404, detail="Index de ligne non valide")

    # Supprimer la ligne spécifiée
    df = df.drop(index=index).reset_index(drop=True)
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne supprimée avec succès"}

# Nouveau endpoint pour récupérer les noms de toutes les tables (feuilles) dans le fichier Excel
@app.get("/api/v1/excel_tables")
def get_excel_tables():
    try:
        # Charger le classeur Excel
        book = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True)
        # Récupérer les noms de toutes les feuilles
        sheet_names = book.sheetnames
        return {"tables": sheet_names}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Fichier Excel introuvable")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la lecture du fichier Excel: {str(e)}")

# Réception-Rame
@app.post("/api/v1/reception_rame")
def add_reception_rame(row: ReceptionRameModel):
    sheet_name = "Réception-Rame"
    df = read_excel_data(sheet_name)
    
    # Ajouter une nouvelle ligne
    new_row = pd.DataFrame([row.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne ajoutée avec succès à Réception-Rame"}

@app.put("/api/v1/reception_rame/{index}")
def update_reception_rame(index: int, row: ReceptionRameModel):
    sheet_name = "Réception-Rame"
    df = read_excel_data(sheet_name)
    
    # Vérifier que l'index est valide
    if index < 0 or index >= len(df):
        raise HTTPException(status_code=404, detail="Index de ligne non valide")
    
    # Modifier la ligne spécifiée
    df.loc[index] = row.dict()
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne mise à jour avec succès dans Réception-Rame"}

# Envoi-Rame
@app.post("/api/v1/envoi_rame")
def add_envoi_rame(row: EnvoiRameModel):
    sheet_name = "Envoi-Rame"
    df = read_excel_data(sheet_name)
    
    # Ajouter une nouvelle ligne
    new_row = pd.DataFrame([row.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne ajoutée avec succès à Envoi-Rame"}

@app.put("/api/v1/envoi_rame/{index}")
def update_envoi_rame(index: int, row: EnvoiRameModel):
    sheet_name = "Envoi-Rame"
    df = read_excel_data(sheet_name)
    
    # Vérifier que l'index est valide
    if index < 0 or index >= len(df):
        raise HTTPException(status_code=404, detail="Index de ligne non valide")
    
    # Modifier la ligne spécifiée
    df.loc[index] = row.dict()
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne mise à jour avec succès dans Envoi-Rame"}

# Réception-Prestataire-RLA
@app.post("/api/v1/reception_prestataire_rla")
def add_reception_prestataire_rla(row: ReceptionPrestataireRLAModel):
    sheet_name = "Réception-Prestataire-RLA"
    df = read_excel_data(sheet_name)
    
    # Ajouter une nouvelle ligne
    new_row = pd.DataFrame([row.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne ajoutée avec succès à Réception-Prestataire-RLA"}

@app.put("/api/v1/reception_prestataire_rla/{index}")
def update_reception_prestataire_rla(index: int, row: ReceptionPrestataireRLAModel):
    sheet_name = "Réception-Prestataire-RLA"
    df = read_excel_data(sheet_name)
    
    # Vérifier que l'index est valide
    if index < 0 or index >= len(df):
        raise HTTPException(status_code=404, detail="Index de ligne non valide")
    
    # Modifier la ligne spécifiée
    df.loc[index] = row.dict()
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne mise à jour avec succès dans Réception-Prestataire-RLA"}

# Envoi-Prestataire-RLA
@app.post("/api/v1/envoi_prestataire_rla")
def add_envoi_prestataire_rla(row: EnvoiPrestataireRLAModel):
    sheet_name = "Envoi-Prestataire-RLA"
    df = read_excel_data(sheet_name)
    
    # Ajouter une nouvelle ligne
    new_row = pd.DataFrame([row.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne ajoutée avec succès à Envoi-Prestataire-RLA"}

@app.put("/api/v1/envoi_prestataire_rla/{index}")
def update_envoi_prestataire_rla(index: int, row: EnvoiPrestataireRLAModel):
    sheet_name = "Envoi-Prestataire-RLA"
    df = read_excel_data(sheet_name)
    
    # Vérifier que l'index est valide
    if index < 0 or index >= len(df):
        raise HTTPException(status_code=404, detail="Index de ligne non valide")
    
    # Modifier la ligne spécifiée
    df.loc[index] = row.dict()
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne mise à jour avec succès dans Envoi-Prestataire-RLA"}

# Réparation-Module
@app.post("/api/v1/reparation_module")
def add_reparation_module(row: ReparationModuleModel):
    sheet_name = "Réparation-Module"
    df = read_excel_data(sheet_name)
    
    # Ajouter une nouvelle ligne
    new_row = pd.DataFrame([row.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne ajoutée avec succès à Réparation-Module"}

@app.put("/api/v1/reparation_module/{index}")
def update_reparation_module(index: int, row: ReparationModuleModel):
    sheet_name = "Réparation-Module"
    df = read_excel_data(sheet_name)
    
    # Vérifier que l'index est valide
    if index < 0 or index >= len(df):
        raise HTTPException(status_code=404, detail="Index de ligne non valide")
    
    # Modifier la ligne spécifiée
    df.loc[index] = row.dict()
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne mise à jour avec succès dans Réparation-Module"}

# Essai
@app.post("/api/v1/essai")
def add_essai(row: EssaiModel):
    sheet_name = "Essai"
    df = read_excel_data(sheet_name)
    
    # Ajouter une nouvelle ligne
    new_row = pd.DataFrame([row.dict()])
    df = pd.concat([df, new_row], ignore_index=True)
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne ajoutée avec succès à Essai"}

@app.put("/api/v1/essai/{index}")
def update_essai(index: int, row: EssaiModel):
    sheet_name = "Essai"
    df = read_excel_data(sheet_name)
    
    # Vérifier que l'index est valide
    if index < 0 or index >= len(df):
        raise HTTPException(status_code=404, detail="Index de ligne non valide")
    
    # Modifier la ligne spécifiée
    df.loc[index] = row.dict()
    
    write_excel_data(sheet_name, df)
    return {"message": "Ligne mise à jour avec succès dans Essai"}
